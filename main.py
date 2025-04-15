import argparse
import json
import sys
from pathlib import Path

import openpyxl
import yaml


def load_config(config_path):
    with open(config_path, 'r', encoding='utf-8') as f:
        return yaml.safe_load(f)

def extract_table_records(ws, table_name, columns_map):
    # openpyxlのテーブル取得
    table = None
    for t in ws._tables.values():
        if t.name == table_name:
            table = t
            break
    if table is None:
        raise ValueError(f"テーブルが見つかりません: {table_name}")
    # Table.refはstr型（例: 'A1:C4'）なのでrange_boundariesで座標取得
    min_col, min_row, max_col, max_row = openpyxl.utils.range_boundaries(table.ref)
    # ヘッダ行取得
    header_row = ws.iter_rows(min_row=min_row, max_row=min_row, min_col=min_col, max_col=max_col, values_only=True)
    headers = next(header_row)
    col_idx_map = {h: i for i, h in enumerate(headers)}
    records = []
    for row in ws.iter_rows(min_row=min_row+1, max_row=max_row, min_col=min_col, max_col=max_col, values_only=True):
        rec = {}
        for excel_col, out_key in columns_map.items():
            if excel_col not in col_idx_map:
                raise ValueError(f"テーブル列が見つかりません: {excel_col}")
            rec[out_key] = row[col_idx_map[excel_col]]
        records.append(rec)
    return records

def extract_range_records(ws, cell_range, columns_map, include_empty_row=False):
    # セル範囲からデータ部分のみ抽出（ヘッダ行含まない前提）
    min_col, min_row, max_col, max_row = openpyxl.utils.range_boundaries(cell_range)
    records = []
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col, values_only=True):
        rec = {}
        for col_pos_str, out_key in columns_map.items():
            col_pos = int(col_pos_str)
            idx = col_pos - 1  # 1始まり→0始まり
            rec[out_key] = row[idx]
        # すべての値がNoneの場合はスキップ（デフォルト）
        if not include_empty_row and all(v is None for v in rec.values()):
            continue
        records.append(rec)
    return records

def get_excel_values(excel_path, value_specs, include_empty_range_row=False):
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    results = {}
    sheetnames = wb.sheetnames
    for spec in value_specs:
        if 'table' in spec:
            # テーブル指定
            table_name = spec['table']
            columns_map = spec['columns']
            # テーブルは全シートから探索
            found = False
            for ws in wb.worksheets:
                if table_name in ws.tables:
                    records = extract_table_records(ws, table_name, columns_map)
                    results[spec.get('name', table_name)] = records
                    found = True
                    break
            if not found:
                raise ValueError(f"テーブルが見つかりません: {table_name}")
        elif 'range' in spec:
            # 範囲指定
            range_str = spec['range']
            columns_map = spec['columns']
            if '!' in range_str:
                sheet_name, cell_range = range_str.split('!', 1)
                ws = wb[sheet_name]
            else:
                ws = wb.active
                cell_range = range_str
            records = extract_range_records(ws, cell_range, columns_map, include_empty_row=include_empty_range_row)
            results[spec.get('name', range_str)] = records
        elif 'named_cell' in spec:
            # 名前付きセル優先
            name = spec['named_cell']
            dn = wb.defined_names.get(name)
            if dn is None:
                raise ValueError(f"名前付きセルが見つかりません: {name}")
            # 参照先（シート名, セルアドレス）を取得
            dest = list(dn.destinations)
            if not dest:
                raise ValueError(f"名前付きセルの参照先が不正です: {name}")
            sheet_name, cell_addr = dest[0]
            sheet = wb[sheet_name]
            value = sheet[cell_addr].value
            results[spec['name']] = value
        else:
            sheet_spec = spec['sheet']
            if isinstance(sheet_spec, str) and sheet_spec.startswith('*'):
                # シート名が'*N'形式の場合、左からN番目のシートを選択
                try:
                    idx = int(sheet_spec[1:]) - 1
                    if idx < 0 or idx >= len(sheetnames):
                        raise IndexError
                    sheet = wb[sheetnames[idx]]
                except (ValueError, IndexError):
                    raise ValueError(f"シート指定が不正です: {sheet_spec} (シート数: {len(sheetnames)})")
            else:
                sheet = wb[sheet_spec]
            value = sheet[spec['cell']].value
            results[spec['name']] = value
    return results

def main():
    parser = argparse.ArgumentParser(description='Excel値取得ツール')
    parser.add_argument('excel', nargs='?', help='Excelファイルパス（コマンドライン優先）')
    parser.add_argument('-c', '--config', default='config.yaml', help='設定ファイルパス')
    parser.add_argument('-o', '--output', help='出力ファイルパス（未指定時は標準出力）')
    parser.add_argument('--include-empty-range-row', action='store_true', help='range指定で全列がNoneの行も出力に含める')
    args = parser.parse_args()

    config = load_config(args.config)
    excel_file = args.excel if args.excel else config.get('excel_file')
    value_specs = config['values']

    if not excel_file:
        print('Excelファイルが指定されていません（コマンドラインまたは設定ファイルで指定してください）', file=sys.stderr)
        sys.exit(1)
    if not Path(excel_file).exists():
        print(f'Excelファイルが見つかりません: {excel_file}', file=sys.stderr)
        sys.exit(1)

    results = get_excel_values(excel_file, value_specs, include_empty_range_row=args.include_empty_range_row)

    if args.output:
        with open(args.output, 'w', encoding='utf-8') as f:
            json.dump(results, f, ensure_ascii=False, indent=2)
        print(f'出力完了: {args.output}')
    else:
        json.dump(results, sys.stdout, ensure_ascii=False, indent=2)
        print()  # 改行

if __name__ == '__main__':
    main()
