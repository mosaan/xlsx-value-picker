import json
import os
import sys
from pathlib import Path

import openpyxl
import yaml  # Import yaml explicitly for the test

sys.path.insert(0, str(Path(__file__).parent.parent / "src"))
import subprocess

# 必要な関数をインポート
from xlsx_value_picker.config import get_excel_values
from xlsx_value_picker.config_loader import ConfigModel, OutputFormat


def create_sample_excel(path):
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Sheet1"
    ws1["A1"] = 123
    ws1["B2"] = "=SUM(1, 2)"
    ws2 = wb.create_sheet("Sheet2")
    ws2["C3"] = "abc"
    wb.save(path)


def create_sample_excel_with_table(path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["ID", "Name", "Score"])
    ws.append([1, "Alice", 90])
    ws.append([2, "Bob", 80])
    ws.append([3, "Carol", 70])
    from openpyxl.worksheet.table import Table, TableStyleInfo

    tab = Table(displayName="Table1", ref="A1:C4")
    style = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    tab.tableStyleInfo = style
    ws.add_table(tab)
    wb.save(path)


def create_sample_excel_with_range(path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["dummy1", "dummy2", "dummy3"])  # ヘッダ行（使わない）
    ws.append([10, 20, 30])
    ws.append([11, 21, 31])
    ws.append([12, 22, 32])
    wb.save(path)


def create_sample_excel_with_empty_rows(path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["col1", "col2", "col3"])  # ヘッダ行
    ws.append([1, 2, 3])
    ws.append([None, None, None])  # 全列Noneの行
    ws.append([4, 5, 6])
    wb.save(path)


def test_get_excel_values(tmp_path):
    excel_path = tmp_path / "sample.xlsx"
    create_sample_excel(excel_path)
    value_specs = [
        {"sheet": "Sheet1", "cell": "A1", "name": "value1"},
        {"sheet": "Sheet1", "cell": "B2", "name": "value2"},
        {"sheet": "Sheet2", "cell": "C3", "name": "value3"},
    ]
    result = get_excel_values(str(excel_path), value_specs)
    assert result["value1"] == 123
    # openpyxlの仕様上、式セルの値はNoneになる
    assert result["value2"] is None
    assert result["value3"] == "abc"


def test_get_excel_values_sheet_index(tmp_path):
    excel_path = tmp_path / "sample.xlsx"
    create_sample_excel(excel_path)
    value_specs = [
        {"sheet": "*1", "cell": "A1", "name": "value1"},  # Sheet1
        {"sheet": "*2", "cell": "C3", "name": "value2"},  # Sheet2
    ]
    result = get_excel_values(str(excel_path), value_specs)
    assert result["value1"] == 123
    assert result["value2"] == "abc"


def test_main_output_file_and_stdout(tmp_path):
    # Excelファイルと設定ファイルを作成
    excel_path = tmp_path / "sample.xlsx"
    create_sample_excel(excel_path)
    config_path = tmp_path / "config.yaml"
    config = f"""\
excel_file: {excel_path}
values:
  - sheet: Sheet1
    cell: A1
    name: value1
  - sheet: Sheet2
    cell: C3
    name: value2
"""
    config_path.write_text(config, encoding="utf-8")
    # 出力ファイル指定
    output_path = tmp_path / "result.json"
    print("config_path:", config_path.resolve())
    print("output_path:", output_path.resolve())
    print("tmp_path files before:", list(tmp_path.iterdir()))

    env = os.environ.copy()
    env["PYTHONPATH"] = str(Path(__file__).parent.parent / "src")
    # 引数を修正：EXCEL_FILEを最後のオプションの後に追加
    result = subprocess.run(
        [
            "python",
            "-m",
            "xlsx_value_picker",
            "--config",
            str(config_path.resolve()),
            "--output",
            str(output_path.resolve()),
            str(excel_path.resolve()),  # EXCEL_FILEを位置引数として追加
        ],
        cwd=tmp_path,
        capture_output=True,
        encoding="utf-8",
        env=env,
    )

    print("stdout:", result.stdout)
    print("stderr:", result.stderr)
    print("tmp_path files after:", list(tmp_path.iterdir()))
    print("returncode:", result.returncode)
    assert output_path.exists()
    with open(output_path, encoding="utf-8") as f:
        data = json.load(f)
    assert data["value1"] == 123
    assert data["value2"] == "abc"
    # 標準出力（--output未指定）
    result = subprocess.run(
        [
            "python", 
            "-m", 
            "xlsx_value_picker", 
            "--config", 
            str(config_path.resolve()),
            str(excel_path.resolve()),  # EXCEL_FILEを位置引数として追加
        ],
        cwd=tmp_path,
        capture_output=True,
        encoding="utf-8",
        env=env,
    )
    print("stdout (no output):", result.stdout)
    print("stderr (no output):", result.stderr)
    print("returncode (no output):", result.returncode)
    # 出力をJSONとしてパースし値を比較
    stdout_json = json.loads(result.stdout)
    assert stdout_json["value1"] == 123
    assert stdout_json["value2"] == "abc"


def test_get_excel_values_named_cell(tmp_path):
    excel_path = tmp_path / "sample.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = 999
    # 名前付きセルを新APIで追加（openpyxl 3.1以降推奨）
    from openpyxl.workbook.defined_name import DefinedName

    dn = DefinedName("MY_CELL", attr_text=f"'{ws.title}'!$A$1")
    wb.defined_names.add(dn)
    wb.save(excel_path)
    value_specs = [
        {"named_cell": "MY_CELL", "name": "foo"},
    ]
    result = get_excel_values(str(excel_path), value_specs)
    assert result["foo"] == 999


def test_get_excel_values_table(tmp_path):
    excel_path = tmp_path / "table.xlsx"
    create_sample_excel_with_table(excel_path)
    value_specs = [{"table": "Table1", "columns": {"ID": "id", "Score": "score"}, "name": "table_data"}]
    result = get_excel_values(str(excel_path), value_specs)
    assert result["table_data"] == [
        {"id": 1, "score": 90},
        {"id": 2, "score": 80},
        {"id": 3, "score": 70},
    ]


def test_get_excel_values_range(tmp_path):
    excel_path = tmp_path / "range.xlsx"
    create_sample_excel_with_range(excel_path)
    value_specs = [{"range": "Sheet1!A2:C4", "columns": {"1": "a", "3": "c"}, "name": "range_data"}]
    result = get_excel_values(str(excel_path), value_specs)
    assert result["range_data"] == [
        {"a": 10, "c": 30},
        {"a": 11, "c": 31},
        {"a": 12, "c": 32},
    ]


def test_get_excel_values_range_skip_and_include_empty(tmp_path):
    excel_path = tmp_path / "emptyrow.xlsx"
    create_sample_excel_with_empty_rows(excel_path)
    value_specs = [{"range": "Sheet1!A2:C4", "columns": {"1": "a", "2": "b", "3": "c"}, "name": "range_data"}]
    # デフォルト（空行スキップ）
    result = get_excel_values(str(excel_path), value_specs)
    assert result["range_data"] == [
        {"a": 1, "b": 2, "c": 3},
        {"a": 4, "b": 5, "c": 6},
    ]
    # include_empty_range_row=True で空行も含める
    result2 = get_excel_values(str(excel_path), value_specs, include_empty_range_row=True)
    assert result2["range_data"] == [
        {"a": 1, "b": 2, "c": 3},
        {"a": None, "b": None, "c": None},
        {"a": 4, "b": 5, "c": 6},
    ]


def test_output_format_from_config(tmp_path):
    """Test that the output format specified in the config file is respected"""
    # Create Excel file and config files for different output formats
    excel_path = tmp_path / "sample.xlsx"
    create_sample_excel(excel_path)

    # Create config with YAML output format
    yaml_config_path = tmp_path / "config_yaml.yaml"
    yaml_config = f"""\
excel_file: {excel_path.resolve()}
values:
  - sheet: Sheet1
    cell: A1
    name: value1
  - sheet: Sheet2
    cell: C3
    name: value2
output:
  format: yaml
"""
    yaml_config_path.write_text(yaml_config, encoding="utf-8")

    # Create config with JSON output format
    json_config_path = tmp_path / "config_json.yaml"
    json_config = f"""\
excel_file: {excel_path.resolve()}
values:
  - sheet: Sheet1
    cell: A1
    name: value1
  - sheet: Sheet2
    cell: C3
    name: value2
output:
  format: json
"""
    json_config_path.write_text(json_config, encoding="utf-8")

    env = os.environ.copy()
    env["PYTHONPATH"] = str(Path(__file__).parent.parent / "src")

    # Test YAML output
    yaml_output_path = tmp_path / "result.yaml"
    yaml_result = subprocess.run(
        [
            "python",
            "-m",
            "xlsx_value_picker",
            "--config",
            str(yaml_config_path.resolve()),
            "--output",
            str(yaml_output_path.resolve()),
            str(excel_path.resolve()),  # EXCEL_FILEを位置引数として追加
        ],
        cwd=tmp_path,
        capture_output=True,
        encoding="utf-8",
        env=env,
    )
    
    print("YAML stdout:", yaml_result.stdout)
    print("YAML stderr:", yaml_result.stderr)
    print("YAML returncode:", yaml_result.returncode)
    
    assert yaml_output_path.exists()
    # Verify the file contains valid YAML
    with open(yaml_output_path, encoding="utf-8") as f:
        yaml_data = yaml.safe_load(f)
    assert yaml_data["value1"] == 123
    assert yaml_data["value2"] == "abc"

    # Test JSON output
    json_output_path = tmp_path / "result.json"
    json_result = subprocess.run(
        [
            "python",
            "-m",
            "xlsx_value_picker",
            "--config",
            str(json_config_path.resolve()),
            "--output",
            str(json_output_path.resolve()),
            str(excel_path.resolve()),  # EXCEL_FILEを位置引数として追加
        ],
        cwd=tmp_path,
        capture_output=True,
        encoding="utf-8",
        env=env,
    )
    
    print("JSON stdout:", json_result.stdout)
    print("JSON stderr:", json_result.stderr)
    print("JSON returncode:", json_result.returncode)
    
    assert json_result.returncode == 0, json_result.stderr
    assert json_output_path.exists()
    # Verify the file contains valid JSON
    with open(json_output_path, encoding="utf-8") as f:
        json_data = json.load(f)
    assert json_data["value1"] == 123
    assert json_data["value2"] == "abc"
