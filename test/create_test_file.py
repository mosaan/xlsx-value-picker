#!/usr/bin/env python
"""
テスト用のExcelファイルを作成するシンプルなスクリプト
"""
import sys
from pathlib import Path

import openpyxl
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.utils import quote_sheetname, absolute_coordinate

# 現在のディレクトリを基準にして、親ディレクトリを参照できるようにする
sys.path.append(str(Path(__file__).parent.parent))

# 出力先ファイル
output_file = Path(__file__).parent / "data" / "test.xlsx"

# Excelファイル作成
wb = openpyxl.Workbook()

# デフォルトシートの名前を変更
ws1 = wb.active
ws1.title = "Sheet1"

# シンプルなデータ入力
ws1['A1'] = "名前"
ws1['B1'] = "値"
ws1['A2'] = "項目1"
ws1['B2'] = 100
ws1['A3'] = "項目2"
ws1['B3'] = 200
ws1['A4'] = "項目3"
ws1['B4'] = 300

# 範囲データ用のシートを追加
ws2 = wb.create_sheet(title="RangeData")
ws2['A1'] = "ID"
ws2['B1'] = "名前"
ws2['C1'] = "値"

for i in range(2, 6):
    ws2[f'A{i}'] = i - 1
    ws2[f'B{i}'] = f"テスト{i - 1}"
    ws2[f'C{i}'] = (i - 1) * 100

# 名前定義を追加
ref = f"{quote_sheetname(ws1.title)}!{absolute_coordinate('B2')}"
defn = DefinedName('TestCell', attr_text=ref)
wb.defined_names.add(defn)

# テーブル定義（openpyxlではテーブルの作成は制限がありますが、簡易版として範囲を作成）
ws3 = wb.create_sheet(title="TableData")
ws3['A1'] = "ID"
ws3['B1'] = "Product"
ws3['C1'] = "Price"

for i in range(2, 6):
    ws3[f'A{i}'] = i - 1
    ws3[f'B{i}'] = f"Product{i - 1}"
    ws3[f'C{i}'] = (i - 1) * 1000

# ファイルに保存
output_file.parent.mkdir(exist_ok=True)
wb.save(output_file)
print(f"テスト用Excelファイルを作成しました: {output_file}")