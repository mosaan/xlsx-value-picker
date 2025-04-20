"""
Excelファイル処理機能のテスト
"""

import sys
from pathlib import Path

import openpyxl
import pytest

# テスト対象モジュールへのパスを追加
sys.path.insert(0, str(Path(__file__).parent.parent / "src"))

from xlsx_value_picker.config_loader import ConfigModel, OutputFormat
from xlsx_value_picker.excel_processor import ExcelValueExtractor
from xlsx_value_picker.exceptions import ExcelProcessingError


def create_test_excel(path):
    """テスト用のExcelファイルを作成する"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = 100
    ws["B2"] = 200
    ws["C3"] = "テスト"
    ws["D4"] = None  # 空セル
    ws2 = wb.create_sheet("Sheet2")
    ws2["A1"] = "Sheet2値1"
    wb.save(path)


class TestExcelValueExtractor:
    """ExcelValueExtractorクラスのテスト"""

    @pytest.fixture
    def excel_file(self, tmp_path):
        """テスト用のExcelファイルを作成してパスを返す"""
        file_path = tmp_path / "test.xlsx"
        create_test_excel(file_path)
        return str(file_path)

    def test_init(self, excel_file):
        """初期化が正しく行われることをテスト"""
        # __init__ ではファイルを開かないので workbook は None のはず
        extractor = ExcelValueExtractor(excel_file)
        assert extractor.excel_path.name == "test.xlsx"
        assert extractor.workbook is None # workbook は None であることを確認

    def test_init_nonexistent_file(self):
        """存在しないファイルで初期化してもエラーにならないことをテスト"""
        # __init__ ではファイルを開かないのでエラーにならない
        try:
             ExcelValueExtractor("nonexistent.xlsx")
        except Exception as e:
             pytest.fail(f"初期化時に予期せぬエラーが発生しました: {e}")

    def test_context_manager_success(self, excel_file):
        """コンテキストマネージャが正常に動作することをテスト"""
        with ExcelValueExtractor(excel_file) as extractor:
            assert extractor.workbook is not None # ファイルが開かれている
        assert extractor.workbook is None # ファイルが閉じられている

    def test_context_manager_file_not_found(self):
        """コンテキストマネージャで存在しないファイルを開くとエラーになることをテスト"""
        with pytest.raises(ExcelProcessingError) as excinfo:
            with ExcelValueExtractor("nonexistent.xlsx") as extractor:
                 pass # ここには到達しないはず
        assert "Excelファイルが見つかりません" in str(excinfo.value)

    def test_extract_values(self, excel_file):
        """設定に基づいて値を正しく抽出できることをテスト"""
        # テスト用の設定モデルを作成
        config = ConfigModel(
            fields={"value1": "Sheet1!A1", "value2": "Sheet1!B2", "text": "Sheet1!C3", "sheet2_value": "Sheet2!A1"},
            rules=[],
            output=OutputFormat(format="json"),
        )

        # 値の抽出 (with文を使用)
        with ExcelValueExtractor(excel_file) as extractor:
            result = extractor.extract_values(config)

        # 結果の検証
        assert result == {"value1": 100, "value2": 200, "text": "テスト", "sheet2_value": "Sheet2値1"}

    def test_extract_values_empty_cell(self, excel_file):
        """空セルを含める/含めないオプションが正しく動作することをテスト"""
        # テスト用の設定モデルを作成
        config = ConfigModel(
            fields={"empty_cell": "Sheet1!D4", "non_empty_cell": "Sheet1!A1"},
            rules=[],
            output=OutputFormat(format="json"),
        )

        with ExcelValueExtractor(excel_file) as extractor:
            # デフォルト（空セルを含めない）
            result1 = extractor.extract_values(config)
            # 空セルを含める
            result2 = extractor.extract_values(config, include_empty_cells=True)

        # 結果の検証
        assert result1 == {"non_empty_cell": 100}
        assert result2 == {"empty_cell": None, "non_empty_cell": 100}

    def test_get_cell_value(self, excel_file):
        """_get_cell_valueが正しく値を取得できることをテスト (with文を使用)"""
        with ExcelValueExtractor(excel_file) as extractor:
            # 正常なセル参照
            assert extractor._get_cell_value("Sheet1!A1") == 100
            assert extractor._get_cell_value("Sheet1!C3") == "テスト"
            assert extractor._get_cell_value("Sheet2!A1") == "Sheet2値1"
            assert extractor._get_cell_value("Sheet1!D4") is None # 空セル

    def test_get_cell_value_invalid_sheet(self, excel_file):
        """存在しないシートを参照するとExcelProcessingErrorが発生することをテスト"""
        with ExcelValueExtractor(excel_file) as extractor:
            with pytest.raises(ExcelProcessingError) as excinfo:
                extractor._get_cell_value("NonExistentSheet!A1")
        assert "シートが見つかりません" in str(excinfo.value)

    def test_get_cell_value_invalid_format(self, excel_file):
        """不正なセル参照形式でExcelProcessingErrorが発生することをテスト"""
        with ExcelValueExtractor(excel_file) as extractor:
            with pytest.raises(ExcelProcessingError) as excinfo:
                extractor._get_cell_value("InvalidFormat")
        assert "無効なセル参照形式です" in str(excinfo.value)

    def test_get_cell_value_invalid_cell(self, excel_file):
        """不正なセル位置を参照するとExcelProcessingErrorが発生することをテスト"""
        with ExcelValueExtractor(excel_file) as extractor:
            with pytest.raises(ExcelProcessingError) as excinfo:
                extractor._get_cell_value("Sheet1!InvalidCell")
        assert "無効なセル参照です" in str(excinfo.value)

    def test_close(self, excel_file):
        """closeメソッドがワークブックを閉じることをテスト"""
        with ExcelValueExtractor(excel_file) as extractor:
            assert extractor.workbook is not None
            extractor.close()
            assert extractor.workbook is None
            # 再度閉じてもエラーにならない
            extractor.close()
            assert extractor.workbook is None

        # withブロック終了後も閉じている
        assert extractor.workbook is None

    def test_get_excel_values_function(self, excel_file):
        """get_excel_values 関数が正しく動作することをテスト"""
        from xlsx_value_picker.excel_processor import get_excel_values

        field_mapping = {"val1": "Sheet1!A1", "val2": "Sheet2!A1"}
        result = get_excel_values(excel_file, field_mapping)
        assert result == {"val1": 100, "val2": "Sheet2値1"}

    def test_get_excel_values_function_error(self):
        """get_excel_values 関数でエラーが発生する場合のテスト"""
        from xlsx_value_picker.excel_processor import get_excel_values

        # 存在しないファイル
        with pytest.raises(Exception) as excinfo:
            get_excel_values("nonexistent.xlsx", {"val1": "Sheet1!A1"})
        assert "Excel値の取得中にエラーが発生しました" in str(excinfo.value)
        assert "Excelファイルが見つかりません" in str(excinfo.value.__cause__) # 元の例外を確認

        # 不正なマッピング (一時ファイル作成)
        with pytest.raises(Exception) as excinfo:
             # 一時的な有効なExcelファイルを作成
             temp_excel = Path("temp_valid.xlsx")
             create_test_excel(temp_excel)
             try:
                 get_excel_values(str(temp_excel), {"val1": "InvalidSheet!A1"})
             finally:
                 temp_excel.unlink() # テスト後にファイルを削除
        assert "Excel値の取得中にエラーが発生しました" in str(excinfo.value)
        assert "シートが見つかりません" in str(excinfo.value.__cause__)
