import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.workbook.defined_name import DefinedName
import os


def create_sample_all():
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Sheet1"
    ws1["A1"] = "りんご"
    ws1["B2"] = '=CONCAT("あ","い")'
    ws1["C3"] = "日本語テスト"
    # 名前付きセル
    dn = DefinedName("MY_CELL", attr_text="'Sheet1'!$A$1")
    wb.defined_names.add(dn)
    # 2枚目のシート
    ws2 = wb.create_sheet("Sheet2")
    ws2["A1"] = "ダミー"
    ws2["B2"] = "バナナ"
    # テーブル用データ
    ws3 = wb.create_sheet("TableSheet")
    ws3.append(["商品ID", "商品名", "点数"])
    ws3.append([1, "みかん", 90])
    ws3.append([2, "ぶどう", 80])
    ws3.append([3, "もも", 70])
    tab = Table(displayName="Table1", ref="A1:C4")
    tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    ws3.add_table(tab)
    # 範囲指定用データ
    ws4 = wb.create_sheet("RangeSheet")
    ws4.append(["ダミー1", "ダミー2", "ダミー3"])  # ヘッダ行（使わない）
    ws4.append(["さくらんぼ", "すいか", "メロン"])
    ws4.append([None, None, None])  # 全列Noneの空行
    ws4.append(["いちご", "なし", "パイナップル"])
    ws4.append(["キウイ", "レモン", "マンゴー"])
    wb.save("sample_all.xlsx")
    print("Excelファイル sample_all.xlsx を作成しました。")


def create_sample_all_yaml():
    yaml_text = """
excel_file: sample_all.xlsx
values:
  - sheet: Sheet1
    cell: A1
    name: value1
  - sheet: "*2"
    cell: B2
    name: value2
  - named_cell: MY_CELL
    name: named
  - table: Table1
    columns:
      商品ID: id
      商品名: name
      点数: score
    name: table_data
  - range: RangeSheet!A2:C6
    columns:
      1: a
      3: c
    name: range_data
"""
    with open("sample_all.yaml", "w", encoding="utf-8") as f:
        f.write(yaml_text)
    print("YAMLファイル sample_all.yaml を作成しました。")


def create_sample_jinja2_yaml():
    yaml_text = """
excel_file: sample_all.xlsx
output:
  format: jinja2
  template: |
    # Excel Data Report
    
    ## 基本データ
    - 果物: {{ data.value1 }}
    - 別のシートの値: {{ data.value2 }}
    - 名前付きセル値: {{ data.named }}
    
    ## テーブルデータ
    | 商品ID | 商品名 | 点数 |
    |--------|--------|------|
    {% for item in data.table_data %}| {{ item.id }} | {{ item.name }} | {{ item.score }} |
    {% endfor %}
    
    ## 範囲データ
    {% for item in data.range_data %}- {{ item.a }} to {{ item.c }}
    {% endfor %}

values:
  - sheet: Sheet1
    cell: A1
    name: value1
  - sheet: "*2"
    cell: B2
    name: value2
  - named_cell: MY_CELL
    name: named
  - table: Table1
    columns:
      商品ID: id
      商品名: name
      点数: score
    name: table_data
  - range: RangeSheet!A2:C6
    columns:
      1: a
      3: c
    name: range_data
"""
    with open("sample_jinja2.yaml", "w", encoding="utf-8") as f:
        f.write(yaml_text)
    print("Jinja2テンプレート設定用YAMLファイル sample_jinja2.yaml を作成しました。")


def create_sample_template_file():
    template_content = """# Excel Data Report (External Template)

## 基本データ
- 果物: {{ data.value1 }}
- 別のシートの値: {{ data.value2 }}
- 名前付きセル値: {{ data.named }}

## テーブルデータ
| 商品ID | 商品名 | 点数 |
|--------|--------|------|
{% for item in data.table_data %}| {{ item.id }} | {{ item.name }} | {{ item.score }} |
{% endfor %}

## 範囲データ
{% for item in data.range_data %}- {{ item.a }} to {{ item.c }}
{% endfor %}
"""
    with open("sample_template.j2", "w", encoding="utf-8") as f:
        f.write(template_content)
    print("Jinja2テンプレートファイル sample_template.j2 を作成しました。")

    yaml_text = """
excel_file: sample_all.xlsx
output:
  format: jinja2
  template_file: sample_template.j2

values:
  - sheet: Sheet1
    cell: A1
    name: value1
  - sheet: "*2"
    cell: B2
    name: value2
  - named_cell: MY_CELL
    name: named
  - table: Table1
    columns:
      商品ID: id
      商品名: name
      点数: score
    name: table_data
  - range: RangeSheet!A2:C6
    columns:
      1: a
      3: c
    name: range_data
"""
    with open("sample_template_file.yaml", "w", encoding="utf-8") as f:
        f.write(yaml_text)
    print("外部テンプレートファイル参照用YAMLファイル sample_template_file.yaml を作成しました。")


if __name__ == "__main__":
    create_sample_all()
    create_sample_all_yaml()
    create_sample_jinja2_yaml()
    create_sample_template_file()
