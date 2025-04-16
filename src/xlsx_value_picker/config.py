from pydantic import BaseModel, TypeAdapter, Field
from typing import List, Dict, Union, Any, Optional
import openpyxl
from pathlib import Path

class ValueSpecBase(BaseModel):
    name: str

    def get_value(self, wb: openpyxl.Workbook, include_empty_range_row: bool = False) -> Any:
        raise NotImplementedError("get_value() must be implemented in subclasses")

class SheetValueSpec(ValueSpecBase):
    sheet: str
    cell: str
    name: str

    def get_value(self, wb: openpyxl.Workbook, include_empty_range_row: bool = False) -> Any:
        sheet_spec = self.sheet
        sheetnames = wb.sheetnames
        if isinstance(sheet_spec, str) and sheet_spec.startswith('*'):
            idx = int(sheet_spec[1:]) - 1
            if idx < 0 or idx >= len(sheetnames):
                raise ValueError(f"シート指定が不正です: {sheet_spec} (シート数: {len(sheetnames)})")
            sheet = wb[sheetnames[idx]]
        else:
            sheet = wb[sheet_spec]
        return sheet[self.cell].value

class NamedCellValueSpec(ValueSpecBase):
    named_cell: str
    name: str

    def get_value(self, wb: openpyxl.Workbook, include_empty_range_row: bool = False) -> Any:
        name = self.named_cell
        dn = wb.defined_names.get(name)
        if dn is None:
            raise ValueError(f"名前付きセルが見つかりません: {name}")
        dest = list(dn.destinations)
        if not dest:
            raise ValueError(f"名前付きセルの参照先が不正です: {name}")
        sheet_name, cell_addr = dest[0]
        sheet = wb[sheet_name]
        return sheet[cell_addr].value

class TableValueSpec(ValueSpecBase):
    table: str
    columns: Dict[str, str]
    name: str

    def get_value(self, wb: openpyxl.Workbook, include_empty_range_row: bool = False) -> Any:
        for ws in wb.worksheets:
            if self.table in ws.tables:
                tbl = ws.tables[self.table]
                ref = tbl.ref
                min_col, min_row, max_col, max_row = openpyxl.utils.range_boundaries(ref)
                if min_col is None or max_col is None or min_row is None or max_row is None:
                    raise ValueError(f"テーブル範囲が不正です: {ref}")
                data = []
                headers = {}
                # ヘッダ行を取得
                for header_col_idx in range(int(min_col), int(max_col) + 1):
                    cell_value = ws.cell(row=min_row, column=header_col_idx).value
                    if cell_value is not None:
                        headers[cell_value] = header_col_idx
                # 各行のデータを取得（ヘッダ行はスキップ）
                for row_idx in range(min_row + 1, max_row + 1):
                    row_data = {}
                    for col_name, out_key in self.columns.items():
                        col_idx = headers.get(col_name)
                        if col_idx is not None:
                            value = ws.cell(row=row_idx, column=col_idx).value
                        else:
                            value = None
                        row_data[out_key] = value
                    # すべての値がNoneの場合はスキップ（デフォルト）
                    if not include_empty_range_row and all(v is None for v in row_data.values()):
                        continue
                    data.append(row_data)
                return data
        raise ValueError(f"テーブルが見つかりません: {self.table}")

class RangeValueSpec(ValueSpecBase):
    range: str
    columns: Dict[int, str]
    name: str

    def get_value(self, wb: openpyxl.Workbook, include_empty_range_row: bool = False) -> Any:
        range_str = self.range
        columns_map = self.columns
        if '!' in range_str:
            sheet_name, cell_range = range_str.split('!', 1)
            ws = wb[sheet_name]
        else:
            raise ValueError(f"範囲指定が不正です: {range_str} (シート名を含めて指定してください)")
            # セル範囲からデータ部分のみ抽出（ヘッダ行含まない前提）
        min_col, min_row, max_col, max_row = openpyxl.utils.range_boundaries(cell_range)
        records = []
        for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col, values_only=True):
            rec = {}
            for col_pos_key, out_key in columns_map.items():
                # Convert string or int key to int
                col_pos = int(col_pos_key)
                idx = col_pos - 1  # 1始まり→0始まり
                rec[out_key] = row[idx]
            # すべての値がNoneの場合はスキップ（デフォルト）
            if not include_empty_range_row and all(v is None for v in rec.values()):
                continue
            records.append(rec)
        return records

ValueSpec = Union[SheetValueSpec, NamedCellValueSpec, TableValueSpec, RangeValueSpec]

class OutputFormat(BaseModel):
    format: str = "json"
    template_file: Optional[str] = None
    template: Optional[str] = None

    def model_post_init(self, __context: Any) -> None:
        """Validate that either template_file or template is provided when format is jinja2."""
        if self.format == "jinja2" and not (self.template_file or self.template):
            raise ValueError("Jinja2 output format requires either template_file or template to be specified")
        
        if self.format == "jinja2" and self.template_file and self.template:
            raise ValueError("Cannot specify both template_file and template, choose one")

class Config(BaseModel):
    excel_file: str
    values: List[ValueSpec]
    output: OutputFormat = Field(default_factory=OutputFormat)


def get_excel_values(
    excel_path: Union[str, Path],
    value_specs: List[ValueSpec],
    include_empty_range_row: bool = False
) -> Dict[str, Any]:
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    results = {}
    for spec in value_specs:
        # pydanticモデルでなければ変換
        if not hasattr(spec, '__fields__'):
            spec = TypeAdapter(ValueSpec).validate_python(spec)
        results[spec.name] = spec.get_value(wb, include_empty_range_row=include_empty_range_row)
    return results